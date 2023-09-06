# Aarun Jury 2023-09-05
# Extracts all tables from the AWS IAM Actions, Resources, and Condition Keys for AWS Services
# https://docs.aws.amazon.com/service-authorization/latest/reference/reference_policies_actions-resources-contextkeys.html
# WARNING: At time of writing, there are 386 AWS Services (and counting!). Therefore, this script can take a long time to run.
# The resulting Excel file is also relatively large at 1.2 MB in size.
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

excel_file = 'output_pandas.xlsx'

# Test to see if there is an existing Excel file
try:
    wb = load_workbook(filename=excel_file)
# If not, create one, save it to disk and delete the default sheet
except:
    wb = Workbook()
    wb.save(excel_file)
del wb['Sheet']
wb.close()

# Extract all tables from the wikipage
links_url = "https://docs.aws.amazon.com/service-authorization/latest/reference/reference_policies_actions-resources-contextkeys.html"
base_url = "https://docs.aws.amazon.com/service-authorization/latest/reference/"
service_urls = []
# Send a GET request to the URL
response = requests.get(links_url)

soup = BeautifulSoup(response.content, "html.parser")
# print(soup)
# Find all the links to service pages
# Find the div with class "highlights" that contains the links
highlights_div = soup.find("div", class_="highlights")

# Find all the anchor tags within the div
links = highlights_div.find_all("a")
for link in links:
    service_name = link.text.strip()  # Extract the text of the link
    service_url = link["href"]  # Extract the href attribute
    url = service_url.replace("./", "")
    # print(f"URL: {url}")
    print(f"Adding Service to list of services: {service_name}")
    service_urls.append(url)
print(f"Number of services: {len(service_urls)}")

# For each Amazon AWS Service
for url in service_urls:
    # Construct the full URL
    full_url = f"{base_url}{url}"
    print(f'Scraping: {full_url}')

    # Create a list of DataFrames from the HTML tables
    dfs = pd.read_html(full_url)

    # Create a new sheet for each service
    service_name = url.replace("list_", "").replace(".html", "")
    # Openpyxl complains if a sheet name is longer than 31 characters
    service_name = service_name[:31]
    # print(f"Service name: {service_name}")

    startrow = 0
    with pd.ExcelWriter(excel_file, mode='a', if_sheet_exists='overlay') as writer:
        for df in dfs:
            print(f"Adding table to sheet for: {service_name}")
            df.to_excel(writer, engine="openpyxl", index=False, startrow=startrow, sheet_name=service_name)
            startrow += (df.shape[0] + 2)
        # Add the full_url to the bottom of the sheet
        sheet = writer.sheets[service_name]
        sheet.cell(row=startrow + 1, column=1, value="Full Service URL:")
        sheet.cell(row=startrow + 1, column=2, value=full_url)